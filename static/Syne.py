import openpyxl
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import win32com
import win32com.client as win32
import pandas as pd
import matplotlib.pyplot as plt
import subprocess
import os
import time
import platform
import paramiko as paramiko
import argparse
import shutil
import requests
from datetime import datetime
from django.core.files import File
from pymongo import MongoClient
from mongoengine import Document, StringField, FileField
from mongoengine import connect

def ssh_execute_script(
    client_IP,
    client_name,
    client_password,
    remote_base_path,
    local_folder_path,
    local_bat_file,
):
    # SSH 세션 시작
    ssh = paramiko.SSHClient()

    # 호스트 메소드 정보 입력
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # SSH를 하기 위한 원격 컴퓨터 정보
    ssh.connect(client_IP, username=client_name, password=client_password)

    # SSH를 통해 점검 프로그램이 들어갈 경로를 생성
    create_folder_command = "mkdir C:\\informationSS3"
    stdizn, stdout, stderr = ssh.exec_command(create_folder_command)

    # SFTP 세션 시작
    sftp = ssh.open_sftp()

    # SFTP를 통해 로컬에 있던 점검프로그램을 원격 컴퓨터의 특정 경로에다 삽입
    sftp.put(local_bat_file, "C:\\informationSS3\\Windo.bat")

    # 만들어둔 경로를 통해, 그 경로의 점검 프로그램을 원격에서 실행
    command = "cd /d C:\\informationSS3 && Windo.bat"
    stdin, stdout, stderr = ssh.exec_command(command)

    # 원격 점검에서 생성되는 SeScore3.txt 가 있는지 확인. 없으면 생길 때 까지 5초동안 대기
    remote_file_to_check = "C:\\informationSS3\\W1~82\\Score\\SeScore3.txt"
    while True:
        try:
            sftp.stat(remote_file_to_check)
            break
        except FileNotFoundError:
            print("점검이 진행중입니다...")
            time.sleep(5)

    # 원격 점검프로그램에서 생성되는 디렉토리 이름들을 구조체로 선언
    subdirectories = ["action", "bad", "good", "log", "Score"]

    # 위에서 선언한 구조체 하나하나를 subdir로 반복해서 아래 함수로 사용
    # remote_base_path = 'C:\\informationSS3\\W1~82' 가 큰 경로라 하면
    # os.path.join(remote_base_path, subdir) 는
    # C:\\informationSS3\\W1~82\\action ...  C:\\informationSS3\\W1~82\\bad..
    # 등으로 반환되는 방식

    for subdir in subdirectories:
        remote_folder_path = os.path.join(remote_base_path, subdir)
        local_subdir_path = os.path.join(local_folder_path, subdir)

        # 로컬에 넣을 경로가 없다면 경로 생성
        if not os.path.exists(local_subdir_path):
            os.makedirs(local_subdir_path)

        # 원격 컴퓨터의 생성파일들을 리스트화 시킴.
        # 만약 오류가 생기면, 생성파일의 이름을 부르고 오류가 났다고 출력
        # 출력이 끝나면 계속 하던거 계속하도록함

        try:
            remote_files = sftp.listdir(remote_folder_path)
        except IOError as e:
            print(f"Error: {e}. Skipping {remote_folder_path}.")
            continue

        # 원격의 각 파일들을 아까 리스트화 된 것을 통해
        # 로컬 폴더에다가 복사해서 가져옴

        for remote_file in remote_files:
            remote_file_path = os.path.join(remote_folder_path, remote_file)
            local_file_path = os.path.join(local_subdir_path, remote_file)
            sftp.get(remote_file_path, local_file_path)

    # report.txt 파일은 따로 복사해서 가져옴
    sftp.get(
        os.path.join(remote_base_path, "report.txt"),
        os.path.join(local_folder_path, "report.txt"),
    )

    # SFTP 세션 종료
    sftp.close()

    # 원격에서 만든 C:\\informationSS3" 이 폴더를 삭제
    # 그럼 원격에서 아무런 흔적이 남지 않을 것임
    # 먼저 지금 실행경로가 C드라이브안의 폴더이기 떄문에, 경로를 벗어나야함.
    # 그렇지 않으면 사용중일라고 안될 우려가 있음

    command1 = "cd C:\\"
    stdin, stdout, stderr = ssh.exec_command(command1)
    delete_command1 = "rd /s /q C:\\informationSS3"
    stdin, stdout, stderr = ssh.exec_command(delete_command1)

    # SSH 세션 종
    ssh.close()


# 함수로 정의한 SSH / SFTP 사용
# 그리고 그에 대한 정보

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("client_IP", help="원격 컴퓨터 IP 정보")
    parser.add_argument("client_name", help="관리자 아이디")
    parser.add_argument("client_password", help="관리자 비밀번호")
    args = parser.parse_args()

    # 현재 시간과 날짜.
    now = datetime.now()
    # 폴더명으로 사용할 형식을 지정.. EX): "YYYY-MM-DD_HH-MM-SS"
    folder_name0 = now.strftime("%Y-%m-%d_%H-%M-%S")
    # 새로운 폴더를 만들고자 하는 경로

    Media_path = os.environ.get("FOLDER_PATH")
    # Media_path = (
    #     r"C:\Users\VinoMatinee\Desktop\projectCTF-main\media\syne" + folder_name0
    # )
    # 폴더가 존재하지 않을 경우에만 폴더를 생성.
    try:
        if not os.path.exists(Media_path):
            os.makedirs(Media_path)
        if not os.path.exists(Media_path + "\\img"):
            os.makedirs(Media_path + "\\img")
        if not os.path.exists(Media_path + "\\txt"):
            os.makedirs(Media_path + "\\txt")
        if not os.path.exists(Media_path + "\\Solution"):
            os.makedirs(Media_path + "\\Solution")
        print("폴더 생성 완료...")
    except:
        print("사용자 폴더 생성 실패")
        print("조치를 취하십시오.")

    client_IP = args.client_IP
    client_name = args.client_name
    client_password = args.client_password
    remote_base_path = "C:\\informationSS3\\W1~82"  # 원격 컴퓨터 경로
    local_folder_path = Media_path + "\\txt"  # 원격 컴퓨터에서 복사한 파일을 옮길 로컬 경로
    local_bat_file = "static\\Windo.bat"  # 원격 컴퓨터에 넣을 점검 파일 경로

    ssh_execute_script(
        client_IP,
        client_name,
        client_password,
        remote_base_path,
        local_folder_path,
        local_bat_file,
    )


# 원격에서 가져온 파일들 중에 점수가 적혀있는 텍스트 정보
# 각각 오픈에서 읽어서 객체선언

with open(local_folder_path + r"\\Score\\AScore.txt", "r", encoding="ANSI") as f:
    Ascore = int(f.read())
AscorePer = round((Ascore / 147) * 100, 2)

with open(local_folder_path + r"\\Score\\SScore.txt", "r", encoding="ANSI") as f:
    Sscore = int(f.read())
SscorePer = round((Sscore / 348) * 100, 2)

with open(local_folder_path + r"\\Score\\PScore.txt", "r", encoding="ANSI") as f:
    Pscore = int(f.read())
PscorePer = round((Pscore / 9) * 100, 2)

with open(local_folder_path + r"\\Score\\LScore.txt", "r", encoding="ANSI") as f:
    Lscore = int(f.read())
LscorePer = round((Lscore / 27) * 100, 2)

with open(local_folder_path + r"\\Score\\SeScore.txt", "r", encoding="ANSI") as f:
    Sescore = int(f.read())
SescorePer = round((Sescore / 168) * 100, 2)

# 읽어올 report 파일 경로를 선언
file_path = local_folder_path + r"\\report.txt"

# 텍스트의 번호 항목에 맞게 변수를 선언하기 위해 공통적으로 있는 특징
# [W-?] 이 시작된다는 점을 이용하기 위한 변수
start_marker = "[W-"
end_marker = "]"

# [W-01]부터 [W-81]까지의 변수를 생성할 것이므로 번호 최대 값 선언
variable_count = 81

# 값을 받아올 구조는 디렉토리 {name : value}
variable_dict = {}

# report 파일의 내용을 읽어드림
with open(file_path, "r") as file:
    text = file.read()

# [W-?] ?는 1부터 시작하므로 햇갈리지 않게 1부터 시작
for i in range(1, variable_count + 1):
    # 시작 인덱스는 읽어드린 내용을 기반으로 [W-i]값부터 시작
    # 여기서 i값의 형태는 01,02...81 같은 형태이므로 zfill(2)를 사용하여 문자열 왼쪽에 10미만의 숫자에 0을 채움
    # 끝의 인덱스는 i+1를 기준으로 함
    start_index = text.find(f"{start_marker}{str(i).zfill(2)}]")
    end_index = text.find(f"{start_marker}{str(i+1).zfill(2)}]")
    # 만약 끝의 인덱스 값이 위와 같은 규칙으로 할당되지 못한경우
    # 값이 하나라고 간주하고 시작 값으로 반환
    if end_index == -1:
        extracted_text = text[start_index:]
    else:
        # 끝의 인덱스 값이 존재한다면, 해당 변수의 값에 시작 인덱스부터 끝 문자열 전까지의 값이 할당
        extracted_text = text[start_index:end_index]
    # 받아온 변수 값을 공백을 제거한 후에,
    # 딕셔너리 형태로 저장하는데, 해당 밸류의 이름은 W{i} 형태로 저장
    # [W-01] 의 value 값 이름 : W1 이런식으로 저장
    variable_dict[f"W{i}"] = extracted_text.strip()

# 엑셀에 사용할 템플릿과 결과물로 나올 엑셀의 경로 선언
filet_path1 = "static\Template.xlsx"
filet_path2 = Media_path + r"\\Solution\\Report.xlsx"

# 시작셀 지정 및 엑셀의 오프셋 값 선언
# column의 오프셋 값은 셀의 열의 위치, row 오프셋은 행의 위치를 담당 (가로 / 세로)
# 이 오프셋값들은 해당 셀의 작업이 끝나고 다음셀로 넘어 갈 때, 이 값들을 참조하여 나아간다
# 지금 0 / 4 이므로 D22에서 시작하면 D(+0)22(+4) 해서 다음값은 D26이 된다.
# 단 , 이 설정은 내가 하기 편하도록 설정한 것이므로 참고
start_cell = "D22"
column_offset = 0
row_offset = 4

# 해당 경로의 엑셀을 로드하여 활성화 한다.
wb = openpyxl.load_workbook(filet_path1)
ws = wb.active

# W{i} 순차적으로 셀에다가 값을 삽입하는 과정을 거침
for i in range(1, 82):
    variable_name = f"W{i}"
    # 시작하는 셀은 전에 선언한 D22
    cell_range = ws[start_cell]
    # 셀의 값은 W{i} 의 해당 value 값으로 저장
    cell_range.value = variable_dict[variable_name]
    # 셀 스타일 형식을 편집한다.
    # 현재 형식은 세로 / 가로 를 기준으로 가운데 정렬 / 자동 줄 바꿈 허용
    cell_range.alignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # 시작 셀의 A1,B2 와 같은 알파벳 중에서 알파벳 문자먼울 가져다가 인덱스로 변환
    start_cell_col = openpyxl.utils.cell.column_index_from_string(start_cell[:1])
    # 시작 셀의 열의 행의 값을 가져오고 오프셋 값과 행의 오프셋 값과 합침
    start_cell_row = int(start_cell[1:]) + row_offset
    # 시작 셀의 인덱스 값을 다시 반환하여 위의 정의한 행의 규칙 값과 합침
    start_cell = openpyxl.utils.cell.get_column_letter(start_cell_col) + str(
        start_cell_row
    )


# 처음 그래프의 형태는 전부 0% 선언
# 값들은 백분율의 형태를 따름
ws["B8"].number_format = "0.00%"
ws["B8"].value = AscorePer / 100

ws["C8"].number_format = "0.00%"
ws["C8"].value = SscorePer / 100

ws["D8"].number_format = "0.00%"
ws["D8"].value = PscorePer / 100

ws["E8"].number_format = "0.00%"
ws["E8"].value = LscorePer / 100

ws["F8"].number_format = "0.00%"
ws["F8"].value = SescorePer / 100

# 엑셀의 해당 이름의 차트를 찾아 객체 선언
# 해당 차트 이름은 Chart1
chart = None
for obj in ws._charts:
    if obj.title == "Chart1":
        chart = obj
        break

if chart is not None:
    # 데이터 갱신 부분은 2열8행 부터 6열 8행의 값들로 선언
    data = Reference(ws, min_col=2, min_row=8, max_col=6, max_row=8)
    # 카테고리는 2열7행 부터 6열 7행 부분들의 값들로 선언
    categories = Reference(ws, min_col=2, min_row=7, max_col=6, max_row=7)
    # 정의한 데이터 / 카테고리는 전에 정의한 변수를 참조하여 설정
    chart.set_categories(categories)
    chart.add_data(data)

# 해당 경로로 저장하고 엑셀 종료
wb.save(filet_path2)
wb.close()

def return_score():
    return AscorePer, SscorePer, PscorePer, LscorePer, SescorePer

print(AscorePer, SscorePer, PscorePer, LscorePer, SescorePer)
print("작업 종료")